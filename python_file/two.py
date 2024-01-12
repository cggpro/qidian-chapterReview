import requests
from bs4 import BeautifulSoup
import openpyxl

# 读取Excel文件
excel_file_path = 'books_reviews.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# 循环每一行，获取bookid，并访问对应网页获取字数
for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
    bookid, bookname, total = row
    url = f'https://www.qidian.com/book/{bookid}/'
    
    # 发送请求获取网页内容
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # 在HTML中找到包含字数的部分
        word_count_tag = soup.find('p', class_='count')
        if word_count_tag:
            word_count_text = word_count_tag.get_text(strip=True)

            # 提取字数信息
            word_count = word_count_text.split('万字')[0]
            print(f'书籍 {bookname} 的总字数为：{word_count}')

            # 将字数写入Excel文件
            sheet.cell(row=index, column=4).value = word_count
        else:
            print(f'未能找到书籍 {bookname} 的总字数信息')
    else:
        print(f'无法获取书籍 {bookname} 的网页内容')

#保存Excel文件
workbook.save('books_reviews.xlsx')